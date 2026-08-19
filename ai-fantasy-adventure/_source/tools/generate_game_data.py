from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

from docx import Document
from openpyxl import load_workbook


ATTR_KEYS = ["S", "O", "CH", "CHA", "Š"]
ATTR_FIELD_KEYS = ["strength", "agility", "intelligence", "charisma", "luck"]


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value.lower())
    ascii_value = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    ascii_value = ascii_value.replace("/", "-")
    return re.sub(r"[^a-z0-9]+", "-", ascii_value).strip("-")


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def parse_signed(value: Any) -> int:
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().replace("−", "-").replace("+", "")
    return int(text)


def docx_table(path: Path, index: int) -> list[list[str]]:
    table = Document(path).tables[index]
    return [[cell.text.strip() for cell in row.cells] for row in table.rows]


def rows_to_dicts(rows: list[list[Any]]) -> list[dict[str, Any]]:
    headers = [str(value).strip() for value in rows[0]]
    return [
        {headers[i]: clean(row[i]) if i < len(row) else None for i in range(len(headers))}
        for row in rows[1:]
        if any(clean(value) is not None for value in row)
    ]


def worksheet_records(ws, header_marker: str | None = None) -> list[dict[str, Any]]:
    rows = [[clean(value) for value in row] for row in ws.iter_rows(values_only=True)]
    header_index = 0
    if header_marker is not None:
        for index, row in enumerate(rows):
            if row and row[0] == header_marker:
                header_index = index
                break
        else:
            raise ValueError(f"Header {header_marker!r} not found in {ws.title}")
    headers = [str(value).strip() if value is not None else f"column_{i}" for i, value in enumerate(rows[header_index])]
    records = []
    for row in rows[header_index + 1 :]:
        if not any(value is not None for value in row):
            continue
        records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return records


def split_csv_text(value: str | None, separator: str = ";") -> list[str]:
    if not value or value == "—":
        return []
    return [part.strip() for part in value.split(separator) if part.strip()]


def build_data(source_dir: Path) -> dict[str, Any]:
    manual_path = source_dir / "AI_Fantasy_Adventure_Manual_v1_0.docx"
    bestiary_path = source_dir / "AI_Fantasy_Adventure_Bestiar_v1_0.xlsx"
    magic_path = source_dir / "AI_Fantasy_Adventure_Magie_a_katalog_kouzel_v1_0.docx"
    equipment_path = source_dir / "AI_Fantasy_Adventure_Katalog_vybaveni_v1_0.docx"

    required = [manual_path, bestiary_path, magic_path, equipment_path]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError("Missing sources: " + ", ".join(missing))

    race_copy = {
        "Člověk": {
            "tagline": "Hrdina, který si svou cestu volí sám.",
            "description": "Lidé jsou přizpůsobiví dobrodruzi bez předem dané silné stránky. Hráč si při tvorbě určí jednu silnou a jednu odlišnou slabou vlastnost.",
            "recommended": ["bojovnik", "hranicar", "kouzelnik", "zlodej", "lecitel", "bard"],
        },
        "Elf": {
            "tagline": "Bystrá mysl, ostrý zrak a cit pro magii.",
            "description": "Elfové vynikají Chytrostí a pozorností. Dokážou rozpoznat vzdálené viditelné detaily, ale v hrubé síle zaostávají.",
            "recommended": ["kouzelnik", "hranicar", "lecitel"],
        },
        "Trpaslík": {
            "tagline": "Pevný jako skála a doma v podzemí.",
            "description": "Trpaslíci staví na Síle, odolném charakteru a hluboké znalosti kamene. Jsou méně obratní, zato výborní v přímých a praktických řešeních.",
            "recommended": ["bojovnik", "hranicar", "lecitel"],
        },
        "Ork": {
            "tagline": "Mohutná síla pod kontrolou odvážného srdce.",
            "description": "Orkové jsou nejsilnější základní rasou. Když jim docházejí Životy, jejich útok zblízka ještě zesílí, ale magie pro ně bývá náročnější.",
            "recommended": ["bojovnik", "hranicar", "bard"],
        },
        "Půlčík": {
            "tagline": "Malá postava, velké štěstí a rychlé ruce.",
            "description": "Půlčíci spoléhají na Obratnost a mimořádné Štěstí. Nejsou silní, ale jednou za dobrodružství dokážou neúspěch proměnit v těsný úspěch.",
            "recommended": ["zlodej", "bard", "hranicar"],
        },
    }

    class_copy = {
        "Bojovník": {
            "tagline": "Stojí v první linii a chrání družinu.",
            "description": "Bojovník řeší nebezpečí silou, odvahou a spolehlivou výzbrojí. Jednou za souboj může převzít útok za blízkého spojence.",
            "style": "Přímý boj, ochrana spojenců a široký výběr zbraní a zbrojí.",
        },
        "Hraničář": {
            "tagline": "Stopař, lučištník a průvodce divočinou.",
            "description": "Hraničář spojuje Sílu s Obratností. Umí automaticky určit směr použitelných stop a dobře se uplatní při průzkumu i střelbě.",
            "style": "Průzkum, boj na dálku, pohyb v přírodě a chytrá příprava.",
        },
        "Kouzelník": {
            "tagline": "Ovládá největší škálu magie.",
            "description": "Kouzelník začíná s vysokou Chytrostí, nejlepšími magickými koeficienty a pěti známými kouzly. Magii vždy rozpozná, ale nemusí znát její přesný účel.",
            "style": "Kouzla, znalosti, řešení hádanek a silné magické útoky.",
        },
        "Zloděj": {
            "tagline": "Tichý, obratný a připravený na zámky i pasti.",
            "description": "Zloděj vyniká Obratností a Štěstím. Běžné plížení zvládá automaticky a díky nástrojům dokáže otevřít cestu tam, kde ostatní neuspějí.",
            "style": "Plížení, zámky, pasti, přesné útoky a chytrá řešení.",
        },
        "Léčitel": {
            "tagline": "Drží družinu na nohou a zahání temnotu.",
            "description": "Léčitel kombinuje Chytrost a Charisma, začíná se čtyřmi kouzly a má nejlepší koeficient Léčení. První pomoc může použít i mimo magický limit.",
            "style": "Léčení, ochrana, očista, podpora a klidné vedení skupiny.",
        },
        "Bard": {
            "tagline": "Příběhy, hudba a správná slova ve správnou chvíli.",
            "description": "Bard spoléhá na Charisma a Obratnost. Jednou ve významné scéně nebo souboji dokáže sobě či spojenci zlepšit bonus hodu o jeden stupeň.",
            "style": "Sociální scény, podpora družiny, hudba a všestrannost.",
        },
    }

    race_rows = rows_to_dicts(docx_table(manual_path, 5))
    race_ability_rows = rows_to_dicts(docx_table(manual_path, 6))
    class_rows = rows_to_dicts(docx_table(manual_path, 7))
    class_ability_rows = rows_to_dicts(docx_table(manual_path, 8))
    hero_matrix = docx_table(manual_path, 9)
    class_loadout_rows = rows_to_dicts(docx_table(equipment_path, 10))

    race_abilities = {}
    for row in race_ability_rows:
        race_name, ability_name = row["Rasa"].split(" – ", 1)
        race_abilities[race_name] = {"name": ability_name, "effect": row["Definitivní účinek"]}

    class_abilities = {}
    for row in class_ability_rows:
        class_name, ability_name = row["Povolání"].split(" – ", 1)
        class_abilities[class_name] = {"name": ability_name, "effect": row["Definitivní účinek"]}

    loadouts = {
        row["Povolání"]: {
            "active": split_csv_text(row["Doporučené aktivní vybavení"]),
            "inventory": split_csv_text(row["Inventář"]),
            "note": row["Poznámka"],
        }
        for row in class_loadout_rows
    }

    races = []
    for row in race_rows:
        name = row["Rasa"]
        races.append(
            {
                "slug": slugify(name),
                "name": name,
                "tagline": race_copy[name]["tagline"],
                "description": race_copy[name]["description"],
                "modifiers": {field: parse_signed(row[key]) for key, field in zip(ATTR_KEYS, ATTR_FIELD_KEYS)},
                "strong": row["Silná"],
                "weak": row["Slabá"],
                "ability": race_abilities[name],
                "recommendedClassSlugs": race_copy[name]["recommended"],
                "assetKey": f"races/{slugify(name)}",
            }
        )

    classes = []
    for row in class_rows:
        name = row["Povolání"]
        classes.append(
            {
                "slug": slugify(name),
                "name": name,
                "tagline": class_copy[name]["tagline"],
                "description": class_copy[name]["description"],
                "playStyle": class_copy[name]["style"],
                "modifiers": {field: parse_signed(row[key]) for key, field in zip(ATTR_KEYS, ATTR_FIELD_KEYS)},
                "ability": class_abilities[name],
                "startingEquipment": loadouts[name],
                "assetKey": f"classes/{slugify(name)}",
            }
        )

    race_lookup = {item["name"]: item for item in races}
    class_lookup = {item["name"]: item for item in classes}
    matrix_races = hero_matrix[0][1:]
    heroes = []
    for row in hero_matrix[1:]:
        class_name = row[0]
        for race_name, cell in zip(matrix_races, row[1:]):
            match = re.fullmatch(r"(\d+)/(\d+)/(\d+)/(\d+)/(\d+)\s*•\s*Ž(\d+)", cell.strip())
            if not match:
                raise ValueError(f"Could not parse hero matrix value: {cell}")
            values = [int(value) for value in match.groups()]
            race = race_lookup[race_name]
            cls = class_lookup[class_name]
            heroes.append(
                {
                    "slug": f"{race['slug']}-{cls['slug']}",
                    "name": f"{race_name} – {class_name}",
                    "raceSlug": race["slug"],
                    "classSlug": cls["slug"],
                    "stats": dict(zip(ATTR_FIELD_KEYS, values[:5])),
                    "hp": values[5],
                    "strong": race["strong"],
                    "weak": race["weak"],
                    "raceAbility": race["ability"],
                    "classAbility": cls["ability"],
                    "startingEquipment": cls["startingEquipment"],
                    "description": f"Tato kombinace spojuje rasové přednosti hrdiny {race_name.lower()} s herním stylem povolání {class_name.lower()}.",
                    "assetKey": f"heroes/{race['slug']}-{cls['slug']}",
                    "futureAssets": ["color", "coloringPage", "characterCard", "levelVariants"],
                }
            )

    equipment_specs = [
        (2, "weapons-melee", "Zbraně na blízko"),
        (3, "weapons-ranged", "Zbraně na dálku"),
        (4, "armor", "Zbroje"),
        (5, "shields", "Štíty"),
        (6, "adventure-gear", "Dobrodružné vybavení"),
        (7, "instruments", "Hudební nástroje"),
        (8, "potions", "Lektvary"),
        (9, "magic-items", "Magické předměty"),
    ]
    equipment = []
    for table_index, category_slug, category_name in equipment_specs:
        for row in rows_to_dicts(docx_table(equipment_path, table_index)):
            name = row["Název"]
            item = {
                "slug": slugify(name),
                "name": name,
                "categorySlug": category_slug,
                "category": category_name,
                "assetKey": f"equipment/{slugify(name)}",
                "futureAssets": ["color", "activity"],
                "fields": {key: value for key, value in row.items() if key != "Název" and value is not None},
            }
            equipment.append(item)

    school_rows = rows_to_dicts(docx_table(magic_path, 1))
    school_visuals = {
        "Obecná magie": {"symbol": "✦", "tone": "violet"},
        "Oheň": {"symbol": "◆", "tone": "ember"},
        "Voda": {"symbol": "◒", "tone": "water"},
        "Vzduch": {"symbol": "ϟ", "tone": "air"},
        "Země": {"symbol": "⬟", "tone": "earth"},
        "Mysl": {"symbol": "◉", "tone": "mind"},
        "Duch": {"symbol": "◇", "tone": "spirit"},
        "Světlá magie": {"symbol": "✺", "tone": "light"},
        "Temná magie": {"symbol": "◐", "tone": "shadow"},
        "Přírodní magie": {"symbol": "❧", "tone": "nature"},
        "Hvězdná magie": {"symbol": "✧", "tone": "star"},
    }
    magic_schools = []
    spells = []
    for index, row in enumerate(school_rows):
        name = row["Škola"]
        school_slug = slugify(name.replace(" magie", "")) if name != "Obecná magie" else "obecna"
        magic_schools.append(
            {
                "slug": school_slug,
                "name": name,
                "description": row["Vymezení"],
                "symbol": school_visuals[name]["symbol"],
                "tone": school_visuals[name]["tone"],
                "assetKey": f"magic/schools/{school_slug}",
            }
        )
        spell_rows = rows_to_dicts(docx_table(magic_path, 9 + index))
        for spell in spell_rows:
            spell_name = spell["Kouzlo"]
            spells.append(
                {
                    "slug": slugify(spell_name),
                    "name": spell_name,
                    "schoolSlug": school_slug,
                    "school": name,
                    "type": spell["Typ"],
                    "minIntelligence": spell["Min. CH"],
                    "minLevel": spell["Min. Level"],
                    "bonus": spell["Bonus"],
                    "target": spell["Cíl"],
                    "duration": spell["Trvání"],
                    "limitations": spell["Omezení"],
                    "effect": spell["Přesný účinek"],
                    "assetKey": f"spells/{school_slug}/{slugify(spell_name)}",
                    "futureAssets": ["effect", "card"],
                }
            )

    wb_values = load_workbook(bestiary_path, data_only=True, read_only=True)
    bestiary_rows = worksheet_records(wb_values["Bestiář"])
    group_rows = worksheet_records(wb_values["Skupiny"])
    defense_rows = worksheet_records(wb_values["Obranné specializace"])
    boss_rows = worksheet_records(wb_values["Bossové"], "Boss")

    group_lookup = {row["Tvor"]: row for row in group_rows if row.get("Tvor")}
    defense_lookup = {row["Tvor"]: row for row in defense_rows if row.get("Tvor")}
    boss_lookup = {row["Boss"]: row for row in boss_rows if row.get("Boss")}

    field_map = {
        "S": "strength",
        "O": "agility",
        "CH": "intelligence",
        "CHA": "charisma",
        "Š": "luck",
        "BŠ": "luckBonus",
        "Životy": "hp",
        "FÚ": "physicalAttack",
        "FO": "physicalDefense",
        "MÚ": "magicAttack",
        "MO": "magicDefense",
    }
    category_slugs = {
        "Zvíře": "zvirata",
        "Člověk/NPC": "lide-npc",
        "Fantasy humanoid": "fantasy-humanoidi",
        "Nemrtvý": "nemrtvi",
        "Nestvůra": "nestvury",
    }
    bestiary = []
    for row in bestiary_rows:
        name = row["Jméno"]
        if not name:
            continue
        group = group_lookup.get(name, {})
        defense = defense_lookup.get(name, {})
        boss = boss_lookup.get(name)
        entry = {
            "slug": slugify(name),
            "name": name,
            "category": row["Kategorie"],
            "categorySlug": category_slugs[row["Kategorie"]],
            "difficulty": row["Obtížnost"],
            "stats": {target: row[source] for source, target in field_map.items()},
            "strong": row["Silná"],
            "weak": row["Slabá"],
            "attackAttribute": row["Útočná vlastnost"],
            "weaponBonus": row["Bonus zbraně"],
            "weapon": row["Zbraň / přirozený útok"],
            "armorBonus": row["Bonus zbroje"],
            "armor": row["Zbroj / přirozená ochrana"],
            "shieldBonus": row["Bonus štítu"],
            "shield": row["Štít"],
            "spellBonus": row["Bonus kouzla"],
            "magicProtection": row["Magická ochrana"],
            "ability": {"name": row["Zvláštní schopnost"], "effect": row["Stručný účinek"]},
            "xp": row["XP"],
            "recommendedGroups": {
                "2": group.get("2 hráči"),
                "3": group.get("3 hráči"),
                "4": group.get("4 hráči"),
            },
            "defenseSpecialization": None,
            "isBoss": boss is not None,
            "boss": None,
            "assetKey": f"bestiary/{slugify(name)}",
            "futureAssets": ["color", "coloringPage", "creatureCard"],
        }
        if defense:
            entry["defenseSpecialization"] = {
                "name": defense.get("Obranná schopnost / kombinace"),
                "effect": defense.get("Přesný mechanický účinek"),
                "resistanceType": defense.get("Typ odolnosti"),
                "weakness": defense.get("Chytré řešení / slabina"),
            }
        if boss:
            entry["boss"] = {
                "adventureType": boss.get("Typ dobrodružství"),
                "signatureAbility": boss.get("Charakteristická schopnost"),
                "weakness": boss.get("Slabina / způsob získání výhody"),
                "alternativeSolutions": boss.get("Možná alternativní řešení"),
                "note": boss.get("Poznámka"),
            }
        bestiary.append(entry)

    equipment_categories = [
        {"slug": slug, "name": name, "count": sum(1 for item in equipment if item["categorySlug"] == slug)}
        for _, slug, name in equipment_specs
    ]
    bestiary_categories = [
        {
            "slug": slug,
            "name": name,
            "count": sum(1 for item in bestiary if item["category"] == name),
        }
        for name, slug in category_slugs.items()
    ]

    return {
        "meta": {
            "title": "AI Fantasy Adventure",
            "version": "1.0",
            "sourceDate": "2026-08-19",
            "sourceLanguage": "cs",
            "sourceFiles": [path.name for path in required],
            "counts": {
                "races": len(races),
                "classes": len(classes),
                "heroes": len(heroes),
                "bestiary": len(bestiary),
                "bosses": sum(1 for item in bestiary if item["isBoss"]),
                "equipment": len(equipment),
                "magicSchools": len(magic_schools),
                "spells": len(spells),
            },
        },
        "races": races,
        "classes": classes,
        "heroes": heroes,
        "bestiaryCategories": bestiary_categories,
        "bestiary": bestiary,
        "equipmentCategories": equipment_categories,
        "equipment": equipment,
        "magicSchools": magic_schools,
        "spells": spells,
        "rules": {
            "attributes": [
                {"key": "S", "name": "Síla", "use": "boj zblízka, zvedání, tlačení, rozbíjení a výdrž"},
                {"key": "O", "name": "Obratnost", "use": "střelba, uhýbání, plížení, lezení a jemná práce"},
                {"key": "CH", "name": "Chytrost", "use": "hádanky, magie, léčení, pátrání a znalosti"},
                {"key": "CHA", "name": "Charisma", "use": "přesvědčování, uklidňování, vedení a vystupování"},
                {"key": "Š", "name": "Štěstí", "use": "náhoda, nečekaná pomoc a riskantní pokus"},
            ],
            "virtualDie": "Vaelor před každým hodem tajně vytvoří novou náhodnou permutaci výsledků 1–6. Hráč pouze řekne číslo 1–6 a AI odhalí přiřazený skutečný hod.",
            "hpFormula": "S + O + ⌈Š/2⌉",
            "luckBonusFormula": "⌈Š/3⌉",
            "adventureScenes": ["Háček", "Průzkum", "Komplikace", "Volba", "Vyvrcholení", "Závěr"],
            "leveling": "Postava začíná na Levelu 1 s 0 XP. Každý nový Level přidá +1 do jedné zvolené vlastnosti do přirozeného maxima 20; hratelné jsou Levely 1–99.",
            "simpleMode": "Bez XP a Levelů lze po významném dobrodružství zvýšit nejvýše jednu vlastnost o +1; krátká epizoda nestačí.",
            "combat": "Každý hráč provede v kole jednu hlavní akci. Významný souboj typicky směřuje ke 3–5 kolům a Vaelor nesmí měnit statbloky uprostřed boje.",
            "magic": "Mana se nepoužívá. Kouzlo je nutné skutečně získat; Min. CH a případný Min. Level pouze dovolují se je naučit.",
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--sources", type=Path, default=Path("public/downloads"))
    parser.add_argument("--output", type=Path, default=Path("app/data/game-data.json"))
    args = parser.parse_args()
    payload = build_data(args.sources)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload["meta"]["counts"], ensure_ascii=False))


if __name__ == "__main__":
    main()
